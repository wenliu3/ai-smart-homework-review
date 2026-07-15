"""临时文档查重路由 — 不依赖学生提交记录，上传文件直接查重

支持双维度查重:
  1. 文本维度 (text.py): N-gram + 词级 TF-IDF，检测文字照抄/改写
  2. 图片维度 (image.py): 感知哈希(aHash+dHash)，检测截图复制粘贴

支持任务书/模板过滤:
  老师可额外上传一份 template_file（任务书/起始代码模板），
  比对前自动从每份提交里剔除模板内容（文本和图片两个维度分别剔除）。
"""
import uuid
import time
import io
import os
import tempfile
from collections import OrderedDict
from typing import List, Optional
from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse, Response
from urllib.parse import quote
from sqlalchemy.orm import Session
from ..deps import get_current_user, require_roles
from ..database import get_db
from ..models import User
from ..plagiarism import (
    run_full_check,
    extract_all_from_file,
    valid_char_count,
    file_to_html,
    get_char_ngram_set,
    safe_build_highlight_pdf,
    make_highlight_cache_key,
    PHRASE_NGRAM,
)
from ..crud.plagiarism_suggestion import generate_plagiarism_suggestion
from ..core.exceptions import BadRequestException
from ..core.response import ok
from ..config import settings

router = APIRouter()

# 内存缓存: OrderedDict 实现 LRU，限制最大条目数防止内存无限膨胀
_CACHE: OrderedDict = OrderedDict()
_CACHE_MAX_SIZE = 100  # 缓存最大条目数
_CACHE_TTL = 1800  # 30分钟


def _delete_cache_files(cache_entry):
    """删除缓存对应的临时文件"""
    plagiarism_dir = settings.plagiarism_path
    for f in (cache_entry.get("file_data") or []):
        file_url = f.get("fileUrl", "")
        filename = file_url.replace("/uploads/plagiarism_tmp/", "")
        if filename:
            file_path = plagiarism_dir / filename
            try:
                if file_path.exists():
                    file_path.unlink()
            except Exception:
                pass


def _clean_cache():
    """清理过期的缓存记录，并淘汰超限的最旧条目（LRU），同时删除对应的临时文件"""
    now = time.time()
    expired = [k for k, v in _CACHE.items() if now - v["timestamp"] > _CACHE_TTL]
    for k in expired:
        _delete_cache_files(_CACHE[k])
        del _CACHE[k]
    while len(_CACHE) > _CACHE_MAX_SIZE:
        k, v = _CACHE.popitem(last=False)
        _delete_cache_files(v)

    # 清理过期的标黄 PDF/docx 产物（hl_ 前缀，含临时查重与作业查重），按文件 mtime 判过期
    _plag_dir = settings.plagiarism_path
    if _plag_dir.exists():
        for _f in _plag_dir.iterdir():
            if _f.name.startswith("hl_") and _f.suffix in (".pdf", ".docx"):
                try:
                    if now - _f.stat().st_mtime > _CACHE_TTL:
                        _f.unlink()
                except Exception:
                    pass


def _plagiarism_docx_disk_path(file_url):
    """从临时查重 fileUrl 还原 docx 磁盘路径（/uploads/plagiarism_tmp/x.docx → plagiarism_path/x.docx）"""
    if not file_url:
        return None
    filename = file_url.replace("/uploads/plagiarism_tmp/", "")
    p = settings.plagiarism_path / filename
    return str(p) if p.exists() else None


def _build_highlight_pdf_url(namespace, self_id, other_id, docx_path, snippets):
    """生成标黄 PDF 的访问 URL；docx 为空/无 snippets/生成失败 → 返回 None（前端可回退）"""
    if not docx_path or not snippets:
        return None
    key = make_highlight_cache_key(namespace, self_id, other_id)
    pdf_path = safe_build_highlight_pdf(docx_path, snippets, key, str(settings.plagiarism_path))
    if not pdf_path:
        return None
    return f"/uploads/plagiarism_tmp/{os.path.basename(pdf_path)}"





@router.post("/plagiarism/adhoc-check")
async def adhoc_check(
    files: List[UploadFile] = File(...),
    template_file: Optional[UploadFile] = File(None),
    passRate: Optional[int] = Form(None),
    phraseWeight: Optional[float] = Form(None),
    topicWeight: Optional[float] = Form(None),
    current_user: User = Depends(require_roles("superadmin", "teacher")),
):
    """临时查重 — 上传多个 docx/txt/pdf 文件，内存处理，不落盘不写库

    可选参数 template_file: 上传一份任务书/起始代码模板，比对前自动剔除模板内容。
    同时进行文本、代码、图片三个维度查重，合并返回结果。
    """
    _clean_cache()

    if len(files) < 2:
        raise BadRequestException(10011, "至少需要2份文件才能进行查重比对")

    # ---- 提取模板的文本和图片 ----
    template_text = ""
    template_images = []
    if template_file:
        try:
            tf_ext = "." + (template_file.filename or "").rsplit(".", 1)[-1].lower() if "." in (template_file.filename or "") else ""
            tf_bytes = await template_file.read()
            tf_tmp = os.path.join(tempfile.gettempdir(), f"plagiarism_template_{uuid.uuid4().hex}{tf_ext}")
            with open(tf_tmp, "wb") as tmp:
                tmp.write(tf_bytes)
            template_text, template_images = extract_all_from_file(tf_tmp, tf_ext)
            os.remove(tf_tmp)
        except Exception:
            pass

    # ---- 生成 checkId（提前，用于文件保存命名） ----
    check_id = str(uuid.uuid4())

    # ---- 提取每份提交的文本和图片 ----
    student_data = []
    image_data = []
    skipped = []
    upload_dir = settings.plagiarism_path

    for f in files:
        filename = f.filename or "unknown.docx"
        try:
            ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            content_bytes = await f.read()
            # 写临时文件提取文本/图片/HTML，提取完即删（不再落盘 plagiarism_tmp，
            # 前端对比预览用 contentHtml，不需要原始 docx 的 URL）
            tmp_path = os.path.join(tempfile.gettempdir(), f"plagiarism_{uuid.uuid4().hex}{ext}")
            with open(tmp_path, "wb") as tmp:
                tmp.write(content_bytes)
            try:
                full_text, images = extract_all_from_file(tmp_path, ext)
                html_content = file_to_html(tmp_path, ext)
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

            # 尝试从文件名解析 姓名_学号
            base = filename.rsplit(".", 1)[0]
            parts = base.split("_")
            name = parts[0].strip() if len(parts) >= 1 else base
            stu_id = parts[1].strip() if len(parts) >= 2 else ""

            # 文本查重数据
            if valid_char_count(full_text) < 40:
                skipped.append({"fileName": filename, "reason": "有效字符数过少，可能是空白文件"})
            else:
                student_data.append({
                    "id": len(student_data) + 1,
                    "studentName": name,
                    "studentNumber": stu_id,
                    "content": full_text,
                    "html": html_content,
                    "fileUrl": "",
                    "ext": ext,
                    "fileName": filename,
                })

            # 图片查重数据
            if images:
                image_data.append({
                    "id": len(image_data) + 1,
                    "studentName": name,
                    "studentNumber": stu_id,
                    "images": images,
                })

        except Exception as e:
            skipped.append({"fileName": filename, "reason": f"解析失败: {e}"})

    # ---- 双维度查重 + 合并 ----
    result = run_full_check(
        text_submissions=student_data,
        image_submissions=image_data if len(image_data) >= 2 else None,
        template_text=template_text or None,
        template_images=template_images or None,
        skipped=skipped,
        pass_rate=passRate, phrase_weight=phraseWeight, topic_weight=topicWeight,
    )

    # 存入内存缓存，供下载报告和对比预览用
    _CACHE[check_id] = {"result": result, "file_data": student_data, "timestamp": time.time()}

    return ok({"checkId": check_id, **result})


@router.get("/plagiarism/{check_id}/compare")
def compare_files(
    check_id: str,
    index_a: int,
    index_b: int,
    current_user: User = Depends(require_roles("superadmin", "teacher")),
):
    """对比预览 — 从内存缓存取两份文件的文本和命中片段，不需要保存文档到数据库"""
    _clean_cache()

    cached = _CACHE.get(check_id)
    if not cached:
        raise BadRequestException(10011, "查重结果已过期(超过30分钟)，请重新查重")

    file_data = cached.get("file_data", [])
    file_a = next((f for f in file_data if f["id"] == index_a), None)
    file_b = next((f for f in file_data if f["id"] == index_b), None)

    if not file_a or not file_b:
        raise BadRequestException(10011, "未找到指定的文件")

    # 计算公共 N-gram 片段（全量返回，供前端 DOM 标黄；只取 top10 会漏标命中段落）
    set_a = get_char_ngram_set(file_a["content"], PHRASE_NGRAM)
    set_b = get_char_ngram_set(file_b["content"], PHRASE_NGRAM)
    common = set_a & set_b
    snippets = [g for g in common if len(g) >= PHRASE_NGRAM]

    # 标黄改由前端 HTML + DOM 标黄实现，后端不再生成 PDF（保留字段为 null，前端忽略）
    pdf_url_a = None
    pdf_url_b = None

    return ok({
        "studentA": {"name": file_a["studentName"], "number": file_a["studentNumber"]},
        "studentB": {"name": file_b["studentName"], "number": file_b["studentNumber"]},
        "fileA": {"fileUrl": file_a.get("fileUrl"), "ext": file_a.get("ext", ""), "fileName": file_a.get("fileName", "")},
        "fileB": {"fileUrl": file_b.get("fileUrl"), "ext": file_b.get("ext", ""), "fileName": file_b.get("fileName", "")},
        "pdfUrlA": pdf_url_a,
        "pdfUrlB": pdf_url_b,
        "contentHtmlA": file_a.get("html") or f"<pre>{file_a['content']}</pre>",
        "contentHtmlB": file_b.get("html") or f"<pre>{file_b['content']}</pre>",
        "snippets": snippets,
    })


@router.delete("/plagiarism/{check_id}/files")
def cleanup_check_files(
    check_id: str,
    current_user: User = Depends(require_roles("superadmin", "teacher")),
):
    """页面退出时清理临时查重文件：删除该 checkId 的原始 word + 标黄 PDF + 内存缓存。

    只用于临时查重工具（adhoc）。学生正式提交的 docx 不走此接口，
    其标黄 PDF 靠 _clean_cache 的 TTL 过期清理。
    """
    cached = _CACHE.pop(check_id, None)
    if cached:
        _delete_cache_files(cached)

    # 删除标黄 PDF/docx 中间产物（前缀 hl_{check_id}_）
    plagiarism_dir = settings.plagiarism_path
    prefix = f"hl_{check_id}_"
    removed = 0
    if plagiarism_dir.exists():
        for f in plagiarism_dir.iterdir():
            if f.name.startswith(prefix) and f.suffix in (".pdf", ".docx"):
                try:
                    f.unlink()
                    removed += 1
                except Exception:
                    pass
    return ok({"message": "已清理", "removedFiles": removed})


@router.get("/plagiarism/{check_id}/report")
def download_report(
    check_id: str,
    current_user: User = Depends(require_roles("superadmin", "teacher")),
):
    """下载查重报告 Excel — 从内存缓存取结果，生成后流式返回"""
    _clean_cache()

    cached = _CACHE.get(check_id)
    if not cached:
        raise BadRequestException(10011, "查重结果已过期(超过30分钟)，请重新查重")

    result = cached["result"]
    results = result.get("results", [])
    has_image = result.get("imageResult") is not None

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "查重结果"

    headers = ["排名", "姓名", "学号",
                "片段重合度(%)", "主题相似度(%)", "综合重复率(%)", "文本判定"]
    if has_image:
        headers += ["图片重合度(%)", "疑似复制图片数", "图片判定"]
    headers += ["最相似对象", "对方学号", "疑似抄袭原因"]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for idx, r in enumerate(results, 1):
        row_data = [
            idx,
            r["studentName"],
            r["studentNumber"],
            r.get("phraseRate") if r.get("phraseRate") is not None else "-",
            r.get("topicRate") if r.get("topicRate") is not None else "-",
            r.get("rate") if r.get("rate") is not None else "-",
            r.get("status", "-"),
        ]
        if has_image:
            row_data += [
                r.get("imageRate") if r.get("imageRate") is not None else "-",
                r.get("matchedImageCount", 0),
                r.get("imageStatus", "-"),
            ]
        row_data += [
            r.get("matchName", "-"),
            r.get("matchId", "-"),
            r.get("suspectReason", "") or "合格",
        ]
        ws.append(row_data)
        row = ws[ws.max_row]

        # 文字查重不合格 → 整行标红
        text_suspect = r.get("status") and r["status"] not in ("合格", "-", None)
        # 图片查重不合格 → 整行标黄
        image_suspect = has_image and r.get("imageStatus") and r["imageStatus"] not in ("合格", "-", None)

        if text_suspect:
            for cell in row:
                cell.fill = red_fill
        elif image_suspect:
            for cell in row:
                cell.fill = yellow_fill

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_len

    # 输出到内存流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"查重报告_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/plagiarism/{check_id}/ai-suggestion")
def ai_suggestion(
    check_id: str,
    index: int,
    compare_index: int = None,
    current_user: User = Depends(require_roles("superadmin", "teacher")),
    db: Session = Depends(get_db),
):
    """AI 建议 — 结合查重结果和大模型，针对学生作业提出分析和建议。

    参数:
        index: 学生在查重列表中的 id
        compare_index: 可选，对比模式下对方的 id，传入后生成对比建议
    """
    _clean_cache()

    cached = _CACHE.get(check_id)
    if not cached:
        raise BadRequestException(10011, "查重结果已过期(超过30分钟)，请重新查重")

    file_data = cached.get("file_data", [])
    results = cached.get("result", {}).get("results", [])

    file_a = next((f for f in file_data if f["id"] == index), None)
    if not file_a:
        raise BadRequestException(10011, "未找到指定的文件")

    # 按姓名+学号匹配查重结果
    plag_info = next(
        (r for r in results
         if r["studentName"] == file_a["studentName"] and r["studentNumber"] == file_a["studentNumber"]),
        {},
    )

    compare_name = None
    compare_content = None
    snippets = None

    if compare_index:
        file_b = next((f for f in file_data if f["id"] == compare_index), None)
        if file_b:
            compare_name = file_b["studentName"]
            compare_content = file_b["content"]
            set_a = get_char_ngram_set(file_a["content"], PHRASE_NGRAM)
            set_b = get_char_ngram_set(file_b["content"], PHRASE_NGRAM)
            common = set_a & set_b
            snippets = sorted(
                [g for g in common if len(g) >= PHRASE_NGRAM],
                key=len, reverse=True,
            )[:10]

    try:
        suggestion = generate_plagiarism_suggestion(
            db,
            student_name=file_a["studentName"],
            student_number=file_a["studentNumber"],
            content=file_a["content"],
            plagiarism_info=plag_info,
            compare_name=compare_name,
            compare_content=compare_content,
            snippets=snippets,
        )
    except ValueError as e:
        raise BadRequestException(10011, str(e))
    return ok({"suggestion": suggestion})
